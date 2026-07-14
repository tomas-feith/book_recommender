"""Parse an uploaded reading list into (title, author) entries.

Users can import an existing library -- from Goodreads/StoryGraph exports or a
hand-made list -- as CSV, TSV, TXT, or XLSX. This module only *parses* the file
into normalized entries; matching them to catalog books lives in the service
(it needs the title index). Kept dependency-light and pure so it's easy to test.

Column detection is forgiving: a header row naming title/author is used if
present; otherwise the first column is the title and the second (if any) the
author. Plain-text files are one book per line, optionally ``Title by Author``.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass

TITLE_HEADERS = {"title", "name", "book", "book title", "booktitle", "original title"}
AUTHOR_HEADERS = {"author", "authors", "by", "writer", "author l-f"}


@dataclass(frozen=True)
class LibraryEntry:
    title: str
    author: str = ""

    def label(self) -> str:
        return f"{self.title} — {self.author}" if self.author else self.title


def parse_library(filename: str, raw: bytes) -> list[LibraryEntry]:
    """Parse an uploaded file's bytes into de-duplicated LibraryEntry rows."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xlsm"):
        entries = _rows_to_entries(_read_xlsx(raw))
    elif ext in ("csv", "tsv"):
        entries = _rows_to_entries(_read_delimited(raw, "\t" if ext == "tsv" else ","))
    elif ext == "txt":
        entries = _parse_lines(raw)
    else:  # unknown extension: sniff -- try delimited, fall back to lines
        entries = _rows_to_entries(_read_delimited(raw, ",")) or _parse_lines(raw)
    return _dedupe(entries)


def _decode(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", "replace")


def _read_delimited(raw: bytes, delimiter: str) -> list[list[str]]:
    text = _decode(raw)
    return list(csv.reader(io.StringIO(text), delimiter=delimiter))


def _read_xlsx(raw: bytes) -> list[list[str]]:
    from openpyxl import load_workbook  # optional dep, imported lazily

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = [["" if c is None else str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _detect_columns(header: list[str]) -> tuple[int | None, int | None]:
    title_i = author_i = None
    for i, h in enumerate(header):
        hl = h.strip().lower()
        if title_i is None and hl in TITLE_HEADERS:
            title_i = i
        if author_i is None and hl in AUTHOR_HEADERS:
            author_i = i
    return title_i, author_i


def _rows_to_entries(rows: list[list[str]]) -> list[LibraryEntry]:
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        return []
    header = [str(c).strip() for c in rows[0]]
    title_i, author_i = _detect_columns(header)
    if title_i is not None:
        body = rows[1:]
    else:  # no recognizable header -> first column is the title
        title_i, author_i, body = 0, (1 if len(header) > 1 else None), rows
    out = []
    for r in body:
        title = r[title_i].strip() if title_i < len(r) else ""
        author = r[author_i].strip() if author_i is not None and author_i < len(r) else ""
        if title:
            out.append(LibraryEntry(title, author))
    return out


def _parse_lines(raw: bytes) -> list[LibraryEntry]:
    out = []
    for line in _decode(raw).splitlines():
        line = line.strip()
        if not line:
            continue
        if "\t" in line:  # "Title<TAB>Author"
            title, _, author = line.partition("\t")
            out.append(LibraryEntry(title.strip(), author.strip()))
            continue
        low = line.lower()
        idx = low.rfind(" by ")  # "Title by Author"
        if idx > 0:
            out.append(LibraryEntry(line[:idx].strip(), line[idx + 4 :].strip()))
        else:
            out.append(LibraryEntry(line, ""))
    return out


def _dedupe(entries: list[LibraryEntry]) -> list[LibraryEntry]:
    seen: set[tuple[str, str]] = set()
    out = []
    for e in entries:
        key = (e.title.strip().lower(), e.author.strip().lower())
        if key not in seen:
            seen.add(key)
            out.append(e)
    return out
